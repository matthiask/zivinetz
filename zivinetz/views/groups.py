from collections import defaultdict
from datetime import timedelta

from django.utils.formats import date_format
from django.utils.translation import activate

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from zivinetz.models import Absence, Assignment, Group, GroupAssignment


letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
columns = {}
columns.update({i: c for i, c in enumerate(letters)})
columns.update({i + 26: "A%s" % c for i, c in enumerate(letters)})
columns.update({i + 52: "B%s" % c for i, c in enumerate(letters)})


def c(column, row):
    """Cell name for zero-indexed column/row combination"""
    return "%s%s" % (columns[column], row + 1)


def create_groups_xlsx(day):
    activate("de")

    day = GroupAssignment.objects.monday(day)
    days = [day + timedelta(days=i) for i in range(5)]

    wb = Workbook()
    ws = wb.active

    thin_border = Side(border_style="thin", color="00000000")
    medium_border = Side(border_style="medium", color="00000000")
    font = Font(name="Calibri")

    dark = NamedStyle("dark")
    dark.font = font
    dark.fill = PatternFill("solid", "cccccc")
    dark.border = Border(
        top=thin_border, right=thin_border, bottom=thin_border, left=thin_border
    )
    wb.add_named_style(dark)

    darker = NamedStyle("darker")
    darker.border = Border(top=thin_border, bottom=thin_border)
    darker.font = Font(name="Calibri", bold=True)
    darker.fill = PatternFill("solid", "aaaaaa")
    wb.add_named_style(darker)

    border = NamedStyle("borderThickLeft")
    border.border = Border(
        top=thin_border, right=thin_border, bottom=thin_border, left=medium_border
    )
    border.font = font
    wb.add_named_style(border)

    border = NamedStyle("borderThickBottom")
    border.border = Border(bottom=medium_border)
    border.font = font
    wb.add_named_style(border)

    border = NamedStyle("borderThinLeft")
    border.border = Border(left=thin_border)
    border.font = font
    wb.add_named_style(border)

    border = NamedStyle("borderThinBottom")
    border.border = Border(bottom=thin_border)
    border.font = font
    wb.add_named_style(border)

    borderThin = NamedStyle("borderThin")
    borderThin.border = Border(
        top=thin_border, right=thin_border, bottom=thin_border, left=thin_border
    )
    borderThin.font = font
    wb.add_named_style(borderThin)

    center = Alignment(horizontal="center", vertical="center")
    vertical_text = Alignment(text_rotation=90)

    def day_column(weekday):
        return 2 + 9 * weekday

    def style_row(row, style):
        for i in range(0, day_column(5) + 1):
            ws[c(i, row)].style = style

    def column_width(column, width):
        ws.column_dimensions[columns[column]].width = width

    def row_height(row, height):
        ws.row_dimensions[row + 1].height = height

    ws[c(0, 1)].style = "borderThickBottom"
    ws[c(1, 1)].style = "borderThickBottom"

    for i, cell in enumerate(
        [
            date_format(day, "F y"),
            "Woche %s" % date_format(day, "W"),
            "Auftragsnummer Arbeit",
            "LEITUNG",
            "ZIVIS",
        ]
    ):
        ws[c(0, i + 1)] = cell
        ws[c(day_column(5), i + 1)] = cell

        ws[c(0, i + 1)].style = "borderThinBottom"
        ws[c(1, i + 1)].style = "borderThinBottom"

        if i > 0:
            ws[c(day_column(5), i + 1)].style = "borderThin"

        if i < 2:
            ws[c(0, i + 1)].alignment = center
            ws[c(day_column(5), i + 1)].alignment = center

    column_width(0, 35)
    column_width(1, 15)
    column_width(day_column(5), 35)

    for i, current in enumerate(days):
        ws[c(day_column(i), 0)] = date_format(current, "l")
        ws[c(day_column(i), 1)] = date_format(current, "d.m.y")
        ws[c(day_column(i), 1)].style = "borderThickBottom"
        ws[c(day_column(i), 0)].alignment = center
        ws[c(day_column(i), 1)].alignment = center
        ws.merge_cells("%s:%s" % (c(day_column(i), 0), c(day_column(i + 1) - 1, 0)))
        ws.merge_cells("%s:%s" % (c(day_column(i), 1), c(day_column(i + 1) - 1, 1)))

        ws[c(day_column(i), 2)] = "Absenz"
        for k in range(2, 499):
            ws[c(day_column(i), k)].style = "borderThickLeft"
            ws[c(day_column(i) + 1, k)].style = "borderThin"
        for j in range(1, 9):
            ws[c(day_column(i) + j, 2)] = "%s)" % j
            style = "borderThin" if j % 2 else "dark"
            for k in range(2, 499):
                ws[c(day_column(i) + j, k)].style = style

            ws[c(day_column(i) + j, 2)].alignment = vertical_text
            column_width(day_column(i) + j, 7)
        ws[c(day_column(i), 2)].alignment = vertical_text
        column_width(day_column(i), 7)

    row_height(2, 250)
    row_height(3, 60)
    row_height(4, 60)

    # ZIVIS line
    style_row(5, "darker")

    assignments = defaultdict(list)
    seen_assignments = set()
    for ga in GroupAssignment.objects.filter(week=day).select_related(
        "assignment__drudge__user"
    ):
        assignments[ga.group_id].append(ga.assignment)
        seen_assignments.add(ga.assignment_id)

    free_assignments = (
        Assignment.objects.for_date(day)
        .exclude(pk__in=seen_assignments)
        .select_related("drudge__user")
    )

    absences = defaultdict(dict)
    for absence in Absence.objects.filter(days__overlap=days):
        for day in absence.days:
            absences[absence.assignment_id][day] = absence

    def add_group(row, group_name, assignments):
        ws[c(0, row)] = group_name
        ws[c(day_column(5), row)] = group_name
        style_row(row, "darker")

        # TODO courses (UNA/MSK)

        for assignment in assignments:
            row += 1
            ws[c(0, row)] = assignment.drudge.user.get_full_name()
            ws[c(day_column(5), row)] = assignment.drudge.user.get_full_name()

            ws[c(0, row)].style = "borderThinBottom"
            ws[c(1, row)].style = "borderThinBottom"
            ws[c(day_column(5), row)].style = "borderThinBottom"

            row_height(row, 35)
            if assignment.date_from in days:
                ws[c(1, row)] = "NEU"
            elif assignment.determine_date_until() in days:
                ws[c(1, row)] = "ENDE"
            else:
                ws[c(1, row)] = date_format(assignment.determine_date_until(), "d.m.y")

            for i, current in enumerate(days):
                if current < assignment.date_from:
                    ws[c(day_column(i), row)] = "Vor Beginn"
                elif current > assignment.determine_date_until():
                    ws[c(day_column(i), row)] = "Nach Ende"
                elif current in absences[assignment.id]:
                    ws[c(day_column(i), row)] = absences[assignment.id][
                        current
                    ].pretty_reason()

        # Skip some lines
        for i in range(0, max(3, 6 - len(assignments))):
            row += 1
            row_height(row, 35)

            ws[c(0, row)].style = "borderThinBottom"
            ws[c(1, row)].style = "borderThinBottom"
            ws[c(day_column(5), row)].style = "borderThinBottom"

        row += 1
        return row

    row = 6
    for group in Group.objects.active():
        row = add_group(row, group.name, assignments[group.id])
    row = add_group(row, "Nicht zugeteilt", free_assignments)

    return wb
